import csv
import datetime
from pathlib import Path

import trade_record

from openpyxl import Workbook

import daily_settlement
import common
import position_stock

downloads_path = str(Path.home() / "Downloads")


class Sheet:

    def __init__(self, sheet_name: str):
        self.sheet_name = sheet_name
        self.fields = []
        self.rows = []

    def add_fields(self, fields):
        for f in fields:
            self.fields.append(f)

    def add_rows(self, rows):
        for row in rows:
            self.rows.append(row)


def write_xlsx(file_name, fields, rows):
    print('running write csv function')
    xlsx_file_name = str(Path.home() / "Downloads" / file_name) + '.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(fields)
    for row in rows:
        ws.append(list(row))
    wb.save(xlsx_file_name)


def write_csv(file_name, fields, rows):
    csv_file_name = str(Path.home() / "Downloads" / file_name) + '.csv'
    with open(csv_file_name, 'w') as f:
        write = csv.writer(f)
        write.writerow(fields)
        write.writerows(rows)


def write_sheet_xlsx(file_name, sheets: []):
    xlsx_file_name = str(Path.home() / "Downloads" / file_name) + '.xlsx'
    wb = Workbook()
    for sheet in sheets:
        ws = wb.create_sheet(sheet.sheet_name)
        ws.append(sheet.fields)
        for row in sheet.rows:
            ws.append(list(row))
    wb.save(xlsx_file_name)


def write_trade_record(file_name=None, start_date=None, end_date=None):
    if start_date is None:
        start_date = '2018-01-01'
    if end_date is None:
        end_date = datetime.datetime.today() + datetime.timedelta(days=1)
    fields, rows = trade_record.get_trade_record(start_date, end_date.strftime('%Y-%m-%d'))
    if file_name is None:
        end_date = end_date + datetime.timedelta(days=-1)
        file_name = str('trade_record_' + start_date + '~' + end_date.strftime('%Y-%m-%d'))
    write_xlsx(file_name, fields, rows)
    print('Done')


def write_trade_record_report(file_name=None, start_date=None, end_date=None):
    if start_date is None:
        start_date = '2018-01-01'
    if end_date is None:
        end_date = datetime.datetime.today() + datetime.timedelta(days=1)
    fields, rows, trade_records = trade_record.get_trade_record(start_date, end_date.strftime(common.date_format))

    if file_name is None:
        end_date = end_date + datetime.timedelta(days=-1)
        file_name = str('trade_report' + start_date + '~' + end_date.strftime(common.date_format))

    position_stocks = position_stock.get_position_stock()
    sheets = []
    daily_settlements = daily_settlement.get_daily_settlement(trade_records)

    ds = Sheet('daily_report')
    ds.add_fields(daily_settlement.daily_settlements_field)
    ds.add_rows(daily_settlements)
    sheets.append(ds)

    tr = Sheet('trade_record')
    tr.add_fields(fields)
    tr.add_rows(rows)
    sheets.append(tr)

    ps = Sheet('position_stock')
    ps.add_fields(position_stock.position_stocks_field)
    ps.add_rows(position_stocks)
    sheets.append(ps)

    write_sheet_xlsx(file_name, sheets)
    print('Done')
